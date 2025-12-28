import pygame
import sys
import os
import time
import math
import random
import types
from textwrap import dedent, shorten, wrap
import json
import openpyxl
from scripts.utils import load_font, drawText, load_image, load_images, load_text, draw_alpha, alphabet_converter, scale_pos, Animation
from scripts.ui_elements import Window, Button, DynamicTextBox, InputBox, set_progress, close_window, change_window_pos_and_dim, check_window, update_ui_elements, ui_process_mouse_button_down, ui_process_keyboard_button_down, render_ui, extract_data, add_window_element, remove_window_element

NAMES_FILE_PATH = "config/employee_names.json"
GENERAL_EXCEPTIONS_FILE_PATH = "config/general_exceptions.json"
WEEKLY_EXCEPTIONS_FILE_PATH = 'config/weekly_exceptions/'
VARIATION_NUMBERS_FILE_PATH = "config/variation_numbers.json"
WORK_CODES_FILE_PATH = "config/work_codes.json"
ISO_NUMBERS_FILE_PATH = "config/iso_numbers.json"
ALLOCATION_TABLE_FILE_PATH = "../data/Raw_Allocation_Tables/"

## Adjust these values based on the current year and how the allocation sheets are structured.
CURRENTYEAR = 24
ALLOCATION_TABLE_DIMENSIONS = (26, 12)
ALLOCATION_TABLE_START_CORNER = (4, 9)
EMPLOYEE_NAME_COLUMN = 'B'
EMPLOYEE_ROLE_COLUMN = 'C'
TASK_NAME_ROW = 3

DAYS_LIST = ['Mon', 'Tues', 'Wed', 'Thur', 'Fri', 'Sat', 'Sun']

class Application(object):
  def __init__(self):

    pygame.init()

    pygame.display.set_caption('ffenics_helper')
    self.screen = pygame.display.set_mode((1024, 768), pygame.RESIZABLE)
    self.display = pygame.Surface((1024, 768))
    

    self.clock = pygame.time.Clock()

    self.assets = {

    }

    self.fonts = {
      'title': load_font('Pixeltype.ttf', 60),
			'subtitle': load_font('Pixeltype.ttf', 30),
			'button_default': load_font('Pixeltype.ttf', 30),
      'weekday_button': load_font('Pixeltype.ttf', 40),
      'input_box_default': load_font('Pixeltype.ttf', 40),
      'input_box_small': load_font('Pixeltype.ttf', 30),
      'label': load_font('Pixeltype.ttf', 40),
			'textbox': load_font('Pixeltype.ttf', 20)
    }

    ### Section for the elements that the application uses, UI elements, entities, etc
    self.ui_elements = {
			'windows': [],
			'buttons': []
		}

    self.settings = {
      'scroll_speed': 30
		}

    self.player_inputs = {
			'mouse': [False, False, False]
		}
    
    self.application_title = self.fonts['title'].render('BGEN FFENICS HELPER', False, "cyan")

    name_file = open(NAMES_FILE_PATH, 'r')
    self.employee_names = json.load(name_file)
    name_file.close()

    variation_numbers_file = open(VARIATION_NUMBERS_FILE_PATH, 'r')
    self.bgen_numbers = set(json.load(variation_numbers_file))
    variation_numbers_file.close()
    ## print(self.bgen_numbers)

    work_codes_file = open(WORK_CODES_FILE_PATH, 'r')
    self.work_codes = set(json.load(work_codes_file))
    work_codes_file.close()

    iso_numbers_file = open(ISO_NUMBERS_FILE_PATH, 'r')
    self.iso_numbers = set(json.load(iso_numbers_file))
    iso_numbers_file.close()

    general_exceptions_file = open(GENERAL_EXCEPTIONS_FILE_PATH, 'r')
    self.general_task_names_exceptions = json.load(general_exceptions_file)
    general_exceptions_file.close()

    self.scan_needed = False
    self.scale_lock = [False, False]

    self.error_markers = {
    }

    """
    name_file = open(NAMES_FILE_PATH, 'w')
    json.dump(employee_names, name_file)
    name_file.close()
    """
    '''
    variation_numbers_file = open(VARIATION_NUMBERS_FILE_PATH, 'w')
    json.dump([203, 120, 1212, 12], variation_numbers_file)
    variation_numbers_file.close()
    '''
    """
    general_exceptions_file = open(GENERAL_EXCEPTIONS_FILE_PATH, 'w')
    json.dump(general_task_names_exceptions, general_exceptions_file)
    general_exceptions_file.close()
    """

  def run_engine(self): ## This is the base function of the entire program
    current_scene = ['start_screen', ['30', '06', [1, 0, 0, 0, 0, 0, 0], {'A Millar stores 1 Allocation 5-5-24.xlsm': [[['D12', 'EmployeeNameMissing'], ['D16', 'EmployeeNameMissing'], ['B10', 'InvalidEmployeeName']], [['B12', 'InvalidEmployeeName'], ['B13', 'InvalidEmployeeName']], [], [], [], [], []], 'N DALY Pipefitter Allocation 5-5-24.xlsm': [[], [], [], [], [['Q3', 'InvalidTaskName']], [], []], 'Nights sh 2  Allocation w-e 5-5-24.xlsm': [[], [], [], [['AC3', 'InvalidTaskName']], [], [], []]}]]
    new_scene = ''
    while new_scene != 'quit':
      if current_scene == 'start_screen':
        self.scale_lock = [False, False]
        pygame.draw.rect(self.screen, (0, 0, 0), (0, 0, 10000, 10000))
        new_scene = self.run_startscreen()
      elif current_scene[0] == 'start_screen': ## RESULT FORMAT: ['start_screen', [scanned_month, scanned_week, scanned_days, self.error_markers]]
        self.scale_lock = [False, False]
        pygame.draw.rect(self.screen, (0, 0, 0), (0, 0, 10000, 10000))
        new_scene = self.run_startscreen(current_scene[1])
      elif current_scene[0] == 'run_open_sheet': ## RESULT FORMAT: ['run_open_sheet', element['content'].buttonID.replace('.xlsm', ''), scanned_week_folder, scanned_days]
        self.scale_lock = [True, False]
        pygame.draw.rect(self.screen, (0, 0, 0), (0, 0, 10000, 10000))
        new_scene = self.run_open_sheet(current_scene[1], current_scene[2], current_scene[3])
      else:
        print('INVALID SCENE')
        pygame.quit()
        sys.exit(1)
      
      current_scene = new_scene
  
  def save_lists(self, list_name: str):
    match list_name:
      case 'names':
        name_file = open(NAMES_FILE_PATH, 'w')
        json.dump(self.employee_names, name_file)
        name_file.close()

      case 'bgen_numbers':
        variation_numbers_file = open(VARIATION_NUMBERS_FILE_PATH, 'w')
        json.dump(self.bgen_numbers, variation_numbers_file)
        variation_numbers_file.close()
      
      case 'work_codes':
        work_codes_file = open(WORK_CODES_FILE_PATH, 'w')
        json.dump(self.work_codes, work_codes_file)
        work_codes_file.close()

      case 'iso_numbers':
        iso_numbers_file = open(ISO_NUMBERS_FILE_PATH, 'w')
        json.dump(self.iso_numbers, iso_numbers_file)
        iso_numbers_file.close()
      
      case 'general_exceptions':
        general_exceptions_file = open(GENERAL_EXCEPTIONS_FILE_PATH, 'w')
        json.dump(self.general_task_names_exceptions, general_exceptions_file)
        general_exceptions_file.close()

      case _:
        if list_name.find('.json') != -1:
          pass
        else:
          print('ERROR: INVALID LIST NAME')

  def pop_up_window(self, pop_up_window: Window):
    self.ui_elements['windows'].append(pop_up_window)
      
    while True:
      self.display.fill((0, 0, 0))

      self.display.blit(self.application_title, self.application_title.get_rect(midleft = (100, 100)))

      mouse_on_window = update_ui_elements(self)

            ## RENDERING SECTION ##

      render_ui(self)

      for event in pygame.event.get():

        if event.type == pygame.KEYDOWN:
          ui_process_keyboard_button_down(self, event)

        if event.type == pygame.MOUSEBUTTONDOWN:
          if event.button == 1:
            self.player_inputs['mouse'][0] = True
            ui_event = ui_process_mouse_button_down(self, 1, mouse_on_window)
            if ui_event != None:
              match ui_event['type']:

                case 'textbox_conclusion':

                  for action in ui_event['actions']:	

                    match action[0]:

                      case 'remove_window':
                        
                        self.ui_elements['windows'].remove(action[1])
          if event.button == 2:
            self.player_inputs['mouse'][2] = True
          if event.button == 3:
            self.player_inputs['mouse'][1] = True

        if event.type == pygame.MOUSEWHEEL:
          print(event.y)
          if event.y > 0:
            ui_process_mouse_button_down(self, 4, mouse_on_window)  
          elif event.y < 0:
            ui_process_mouse_button_down(self, 5, mouse_on_window)  

        if event.type == pygame.MOUSEBUTTONUP:
          if event.button == 1:
            self.player_inputs['mouse'][0] = False

            ### I think I'm gonna not do a function for this, it just feels too specialised

            if not mouse_on_window:
              for button in reversed(self.ui_elements['buttons'].copy()):
                if button.get_rect().collidepoint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)) and button.state == 'clicked':
                  match button.buttonID:

                    case _:
                      print('ERROR: INVALID BUTTON ID')
                      assert False
                  break			

            else:
              for window in reversed(self.ui_elements['windows'].copy()):

                for element in reversed(window.elements.copy()):
                  if element['type'] == 'button':
                    if element['content'].get_rect(offset=window.position).collidepoint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)) and element['content'].state == 'clicked':
                      if window.windowID == 'SaveSheetPopUpWindow':
                        print('asd')
                        close_window(self, 'SaveSheetPopUpWindow')
                        return element['content'].buttonID

          if event.button == 2:
            self.player_inputs['mouse'][2] = False
          if event.button == 3:
            self.player_inputs['mouse'][1] = False

        if event.type == pygame.QUIT:
          pygame.quit()
          sys.exit()
      
      self.screen.blit(pygame.transform.scale(self.display, (self.screen.get_width(), self.display.get_height()*self.screen.get_width() / self.display.get_width())), (0, 0))

      pygame.display.update()

  def run_startscreen(self, data= ['', '', [0, 0, 0, 0, 0, 0, 0], {}]):

    def scan_sheets(week_folder, days):

      def check_for_ZZ007(task_cell):
        ## the following code is for dealing with the case where the cell is under the ZZ-007 task, this is mostly for doing checks on whether there are invalid employee names whose rows are entirely blank for some reason
        ## Plus, this code might be able to be reused for when I actually make the summary table
        
        ## first we ignore if the task name if it is blank
        task_name = read_sheet[task_cell].value

        if task_name != None and str(task_name).upper().replace(' ', '') != '':
          task = str(task_name).upper().replace(' ', '')
          ## Next, we check the task to see if it is in any exceptions
          if task in self.general_task_names_exceptions.keys():
            task_exception = self.general_task_names_exceptions[task]
            print(f'{task} replaced with {task_exception}')
            read_sheet[task_cell] = task_exception.upper().replace(' ', '')
            task = task_exception

          elif task in week_specific_task_name_exceptions[sheet_number].keys():
            task_exception = week_specific_task_name_exceptions[sheet_number][task]
            print(f'{task} replaced with {task_exception} (WEEK SPECIFIC)')
            read_sheet[task_cell] = task_exception.upper().replace(' ', '')
            task = task_exception
        
        else: 
          task = ''

        ## Having done that, we then check if the task name is 'ZZ-007'
        ## DISCLAIMER: I am aware that this means that invalid task names can slip through the cracks if there are no actual hours on the task column, but we don't really need that, the employee names are more required

        if task == 'ZZ-007':
          employee_name = read_sheet[f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'].value
          ## print(employee_name)

          if employee_name != None and str(employee_name).strip() != '' and str(employee_name).strip() not in self.employee_names.keys() and j not in ignore_row_list:
            add_error(worksheet_file, sheet_number, f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}', 'InvalidEmployeeName')
            ignore_row_list.append(j) 

      def add_error(file, day, cell, error):
        if file not in self.error_markers.keys():
          self.error_markers[file] = [[],[],[],[],[],[],[]]

        if [cell, error] not in self.error_markers[file][day]: ## Like, I KNOW that this is basically having 2 safeguards against doubling up on errors
                                                               ## NOTE_TO_SELF: this will mess up if you can somehow assign 2 distinct errors to the same cell
          self.error_markers[file][day].append([cell, error])

      print(f'Scanning sheets in file: {week_folder}, with days {days}')

      progress = 0

      print(f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}')

      if not os.path.exists(f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}'):
        print('AllocationTableFileNotFound')
        return 'AllocationTableFileNotFound'
      
      try:
        week_specific_exceptions_file = open(f'{WEEKLY_EXCEPTIONS_FILE_PATH}{week_folder}.json', 'r')
        week_specific_task_name_exceptions = json.load(week_specific_exceptions_file)
        week_specific_exceptions_file.close()
      except FileNotFoundError:
        week_specific_task_name_exceptions = [{},{},{},{},{},{},{}]
        week_specific_exceptions_file = open(f'{WEEKLY_EXCEPTIONS_FILE_PATH}{week_folder}.json', 'x')
        json.dump(week_specific_task_name_exceptions, week_specific_exceptions_file)
        week_specific_exceptions_file.close()

      self.error_markers = {
      }

      for worksheet_file in sorted(os.listdir(f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}')):
        if not worksheet_file.endswith(('.xlsx', '.xlsm')):
          continue

        read_workbook = openpyxl.load_workbook(filename=f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}/{worksheet_file}', read_only=False, data_only=True)
        
        ## the followin
        # print(read_workbook.sheetnames)
        
        if read_workbook.sheetnames != ['Mon', 'Tues', 'Wed', 'Thur', 'Fri', 'Sat', 'Sun', 'Summary', 'Lists']:
          print("Error, Incorrect Sheet Names")
          return ['InvalidWorksheetSheetNames', worksheet_file]

        num_of_days_selected = 0

        for day in days:
          num_of_days_selected += day

        for sheet_number in range(7):
          if days[sheet_number] == 1:
            read_workbook.active = sheet_number
            ## print(worksheet_file + str(sheet_number) + "~###################################################################################################################################################~~")
            read_sheet = read_workbook.active

            ignore_letter_list = [] ## this ensures that whenever we pick up a non number in the table (aka an absence) we don't end up picking it up again.
            ignore_row_list = []
            ignore_column_list = []

            for i in range(ALLOCATION_TABLE_DIMENSIONS[0]):
              for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
                current_cell = f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{ALLOCATION_TABLE_START_CORNER[1] + j}'
                current_cell_value = read_sheet[current_cell].value

                ## print(f"{current_cell_value}|{read_sheet[f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'].value}|{read_sheet[f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}'].value}")
                ## print(ignore_column_list)
                ## print(ignore_row_list)
                ## print(current_cell_value)
                ## print(current_cell_value)

                if current_cell_value != None: ## the cell is not empty

                  ## Checking if the cell just has white space in it
                  if str(current_cell_value).replace(' ', '') == '' and str(current_cell_value) != '':
                    add_error(worksheet_file, sheet_number, current_cell, 'WhitespaceInCell')
                  
                  ## Checking if the cell has a non-number value, usually means that it is an absence
                  elif not str(current_cell_value).replace('.', '').isnumeric():
                    if j not in ignore_letter_list:
                      employee_name = read_sheet[f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'].value
                      ## print(employee_name)
                      if employee_name == None or str(employee_name).replace(' ', '') == '':
                        add_error(worksheet_file, sheet_number, current_cell, 'EmployeeNameMissing')

                      elif str(employee_name).strip() not in self.employee_names.keys() and j not in ignore_row_list:
                        ## print('asd')
                        ## print(f"{worksheet_file} {DAYS_LIST[sheet_number]} {str(employee_name).strip()}")
                        add_error(worksheet_file, sheet_number, f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}', 'InvalidEmployeeName')
                        ignore_row_list.append(j)

                      else:
                        pass
                        ## this section is for adding the absences
                  
                      ignore_letter_list.append(j)  ## See above

                    check_for_ZZ007(f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}')

                  else: ## the cell has a number value, proceed as usual
                    # print(f"{str(current_cell_value).replace('.', '')} is number|| {str(current_cell_value).replace('.', '').isnumeric()}")
                    
                    employee_name = read_sheet[f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'].value
                    ## print(f"{employee_name}|{str(employee_name).strip() not in self.employee_names.keys()}|{j not in ignore_row_list}" )
                    if employee_name == None or str(employee_name).replace(' ', '') == '' and j not in ignore_row_list:
                      add_error(worksheet_file, sheet_number, current_cell, 'EmployeeNameMissing')
                      ignore_row_list.append(j)

                    elif str(employee_name).strip() not in self.employee_names.keys() and j not in ignore_row_list:
                      print(f"{worksheet_file} {DAYS_LIST[sheet_number]} {str(employee_name).strip()}")
                      add_error(worksheet_file, sheet_number, f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}', 'InvalidEmployeeName')
                      ignore_row_list.append(j) 

                    task_name = read_sheet[f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}'].value

                    if task_name == None or str(task_name).upper().replace(' ', '') == '' and i not in ignore_column_list:
                      add_error(worksheet_file, sheet_number, current_cell, 'TaskNameMissing')
                      ignore_column_list.append(i)

                    else:
                      task = str(task_name).upper().replace(' ', '')
                      if task in self.general_task_names_exceptions.keys():
                        task_exception = self.general_task_names_exceptions[task]
                        ## print(f'{task} replaced with {task_exception}')
                        read_sheet[f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}'] = task_exception.upper().replace(' ', '')
                        task = task_exception

                      elif task in week_specific_task_name_exceptions[sheet_number].keys():
                        task_exception = week_specific_task_name_exceptions[sheet_number][task]
                        ## print(f'{task} replaced with {task_exception} (WEEK SPECIFIC)')
                        read_sheet[f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}'] = task_exception.upper().replace(' ', '')
                        task = task_exception

                      if task.find('BGENGRAM') != -1:
                        if task.replace('BGENGRAM', '') == '':
                          return ['MissingBgenGramNumber', worksheet_file, read_workbook.sheetnames[sheet_number], current_cell]
                        
                        elif not task.replace('BGENGRAM', '').isnumeric():
                          ## print(f"{task.replace('BGENGRAM', '')} is not numeric")
                          add_error(worksheet_file, sheet_number, f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}', 'BgenNumberNotAdded')
                          ignore_column_list.append(i)

                        elif (int(task.replace('BGENGRAM', '')) not in self.bgen_numbers) and i not in ignore_column_list:
                          ## print(f"{task.replace('BGENGRAM', '')} not in the list of bgen numbers {int(task.replace('BGENGRAM', '')) not in self.bgen_numbers}")
                          
                          add_error(worksheet_file, sheet_number, f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}', 'BgenNumberNotAdded')
                          ignore_column_list.append(i)

                      elif task not in self.work_codes and task not in self.iso_numbers and i not in ignore_column_list:
                        ## print(f"{task} not in the list of vaild task names {task not in self.work_codes} and {task not in self.iso_numbers}")
                        ## print(f"{worksheet_file} {DAYS_LIST[sheet_number]} {task}")
                        add_error(worksheet_file, sheet_number, f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}', 'InvalidTaskName')
                        ignore_column_list.append(i)
                else:
                  ## read_sheet[f'{current_cell}'] = ''  ## This code was to test if setting a cell to an empty string will cause it's value to be considered None upon future reads, it turns out the answer is yes
                  
                  check_for_ZZ007(f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}')

            progress += 1

          set_progress(self, 'scanSheetLoadingBar', progress/(len(sorted(os.listdir(f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}')))* num_of_days_selected))

          self.display.fill((0, 0, 0))
      
          self.display.blit(self.application_title, self.application_title.get_rect(midleft = (100, 100)))

          mouse_on_window = update_ui_elements(self)

                ## RENDERING SECTION ##

          render_ui(self)

          for event in pygame.event.get():

            if event.type == pygame.QUIT:
              pygame.quit()
              sys.exit()
          
          self.screen.blit(self.display, (0, 0))

          pygame.display.update()

        read_workbook.save(filename=f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}/{worksheet_file}')
        read_workbook.close()
      
      print(self.error_markers)

      return 'ScanCompleted'
    
    self.display = pygame.Surface((1024, 768))

    self.ui_elements = {
      'windows': [Window(self, self.display, 
                         windowID= "FileNameInputWindow", 
                         dimensions= (425, 120), position= (400, 200), 
                         elements= [
                                    {'type': 'label',
                                     'position': (10, 10),
                                     'txt_font': 'label',
                                     'txt_content': 'MONTH:',
                                     'txt_color': 'white'},
                                    {'type': 'label',
                                     'position': (220, 10),
                                     'txt_font': 'label',
                                     'txt_content': 'WEEK:',
                                     'txt_color': 'white'},
                                    {'type': 'input_textbox/default',
                                     'dimensions': (60, 40),
                                     'position': (100, 10),
                                     'txt_content': f'{data[1]}',
                                     'input_boxID': 'MonthInputBox',
                                     'char_limit': 2},
                                    {'type': 'input_textbox/default',
                                     'dimensions': (60, 40),
                                     'position': (300, 10),
                                     'txt_content': f'{data[0]}',
                                     'input_boxID': 'DayInputBox',
                                     'char_limit': 2},
                                    {'type': f'switch/{"off" if data[2][0] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (5, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Mon',
                                    'switchID': 'MondaySwitch'},
                                    {'type': f'switch/{"off" if data[2][1] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (65, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Tue',
                                    'switchID': 'TuesdaySwitch'},
                                    {'type': f'switch/{"off" if data[2][2] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (125, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Wed',
                                    'switchID': 'WednesdaySwitch'},
                                    {'type': f'switch/{"off" if data[2][3] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (185, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Thu',
                                    'switchID': 'ThursdaySwitch'},
                                    {'type': f'switch/{"off" if data[2][4] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (245, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Fri',
                                    'switchID': 'FridaySwitch'},
                                    {'type': f'switch/{"off" if data[2][5] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (305, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Sat',
                                    'switchID': 'SaturdaySwitch'},
                                    {'type': f'switch/{"off" if data[2][6] == 0 else "on"}', 
                                    'dimensions': (55, 40), 
                                    'position': (365, 65),  
                                    'txt_font': 'weekday_button', 
                                    'txt_content': 'Sun',
                                    'switchID': 'SundaySwitch'}
                                     ])],

			'buttons': [Button(self, self.display, (200, 50), (100, 300), txt_align= 'left', txt_content= 'Scan Sheets', buttonID= 'ScanSheetsButton'),
						Button(self, self.display, (200, 50), (100, 400), txt_align= 'left', txt_content= 'Access Sheets', buttonID= 'AccessSheetsButton'),
						Button(self, self.display, (200, 50), (100, 500), txt_align= 'left', txt_content= 'How to Use', buttonID= 'HowToUseButton'),
						Button(self, self.display, (200, 50), (100, 600), txt_align= 'left', txt_content= 'Exit', buttonID= 'ExitButton')]
		}
    
    scanned_week_folder = f"WE.{data[0]}.{data[1]}.{CURRENTYEAR}"
    scanned_days = data[2]
    self.error_markers = data[3]
    
    while True:
      self.display.fill((0, 0, 0))
  
      self.display.blit(self.application_title, self.application_title.get_rect(midleft = (100, 100)))

      mouse_position = scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)

      mouse_on_window = update_ui_elements(self)

      			## RENDERING SECTION ##

      render_ui(self)

      pygame.draw.circle(self.display, 'red', mouse_position, 5, 1)

      ##### EVENT HANDLING SECTION #####	

      for event in pygame.event.get():

        if event.type == pygame.KEYDOWN:
          ui_process_keyboard_button_down(self, event)

        if event.type == pygame.MOUSEBUTTONDOWN:
          if event.button == 1:
            self.player_inputs['mouse'][0] = True
            ui_event = ui_process_mouse_button_down(self, 1, mouse_on_window)
            if ui_event != None:
              match ui_event['type']:

                case 'textbox_conclusion':

                  for action in ui_event['actions']:	

                    match action[0]:

                      case 'remove_window':
                        
                        self.ui_elements['windows'].remove(action[1])
          if event.button == 2:
            self.player_inputs['mouse'][2] = True
          if event.button == 3:
            self.player_inputs['mouse'][1] = True

        if event.type == pygame.MOUSEWHEEL:
          print(event.y)
          if event.y > 0:
            ui_process_mouse_button_down(self, 4, mouse_on_window)  
          elif event.y < 0:
            ui_process_mouse_button_down(self, 5, mouse_on_window)  

        if event.type == pygame.MOUSEBUTTONUP:
          if event.button == 1:
            self.player_inputs['mouse'][0] = False

            ### I think I'm gonna not do a function for this, it just feels too specialised

            if not mouse_on_window:
              for button in reversed(self.ui_elements['buttons'].copy()):
                if button.get_rect().collidepoint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)) and button.state == 'clicked':
                  match button.buttonID:

                    case 'ScanSheetsButton':

                      if check_window(self, 'scanSheetStatusWindow'):
                        close_window(self, 'scanSheetStatusWindow')

                      self.ui_elements['windows'].append(Window(self, self.display, 
                                                                windowID= 'scanSheetStatusWindow', 
                                                                dimensions= (300, 50),
                                                                position= (320, 300),
                                                                elements= [
                                                                  {'type': 'loading_bar/default', 
                                                                   'dimensions': (290, 40),
                                                                   'position': (5, 5),
                                                                   'loading_barID': 'scanSheetLoadingBar'}
                                                                ]))
                      if f"{extract_data(self, 'MonthInputBox')}{extract_data(self, 'DayInputBox')}".upper() == "TEST":
                        scanned_week_folder = "Testing"
                      else:
                        scanned_week_folder = f"WE.{extract_data(self, 'DayInputBox')}.{extract_data(self, 'MonthInputBox')}.{CURRENTYEAR}"
                      
                      scanned_days = [extract_data(self, 'MondaySwitch'),
                                        extract_data(self, 'TuesdaySwitch'),
                                        extract_data(self, 'WednesdaySwitch'),
                                        extract_data(self, 'ThursdaySwitch'),
                                        extract_data(self, 'FridaySwitch'),
                                        extract_data(self, 'SaturdaySwitch'),
                                        extract_data(self, 'SundaySwitch')]

                      scan_result = scan_sheets(scanned_week_folder, scanned_days)
                      
                      match scan_result:
                        case 'AllocationTableFileNotFound':
                          remove_window_element(self, 'scanSheetStatusWindow','scanSheetLoadingBar')
                          add_window_element(self, 'scanSheetStatusWindow', 
                                      {'type': 'label',
                                      'position': (5, 5),
                                      'txt_font': 'label',
                                      'txt_content': 'Allocation Table Folder Not Found',
                                      'txt_color': 'white'},)
                          change_window_pos_and_dim(self, 'scanSheetStatusWindow', dimensions= (500, 50))
                        
                        case 'ScanCompleted':
                          remove_window_element(self, 'scanSheetStatusWindow','scanSheetLoadingBar')
                          add_window_element(self, 'scanSheetStatusWindow', 
                                      {'type': 'label',
                                      'position': (5, 5),
                                      'txt_font': 'label',
                                      'txt_content': 'Allocation Tables Scanned',
                                      'txt_color': 'white'},)
                          change_window_pos_and_dim(self, 'scanSheetStatusWindow', dimensions= (500, 50))
                          self.scan_needed = False

                        case None:
                          pass

                        case _:
                          if type(scan_result) is list:
                            if scan_result[0] == 'InvalidWorksheetSheetNames':
                              remove_window_element(self, 'scanSheetStatusWindow','scanSheetLoadingBar')
                              add_window_element(self, 'scanSheetStatusWindow', 
                                          {'type': 'textbox',
                                          'position': (5, 5),
                                          'dimensions': (490, 40),
                                          'text': f'File {scan_result[1]} Has Invalid Sheet Names',
                                          'color': 'white'},)
                              change_window_pos_and_dim(self, 'scanSheetStatusWindow', dimensions= (500, 50))
                            elif scan_result[0] == 'MissingBgenGramNumber':
                              remove_window_element(self, 'scanSheetStatusWindow','scanSheetLoadingBar')
                              add_window_element(self, 'scanSheetStatusWindow', 
                                          {'type': 'textbox',
                                          'position': (5, 5),
                                          'dimensions': (490, 40),
                                          'text': f'File {scan_result[1]}, Sheet {scan_result[2]}, Cell {scan_result[3]} Has MissingBgenGramNumber',
                                          'color': 'white'},)
                              change_window_pos_and_dim(self, 'scanSheetStatusWindow', dimensions= (500, 50))
                          else:
                            print('ERROR: INVALID SCAN RESULT')
                            assert False

                    case 'AccessSheetsButton':
                      
                      if not self.scan_needed:
                        if check_window(self, 'SheetListWindow'):
                          close_window(self, 'SheetListWindow')
                        else:
                          window_buttons = [{'type': 'textbox',
                                            'position': (5, 5),
                                            'dimensions': (10, 490),
                                            'text': '',
                                            'color': 'white'}]
                          yposition = 5

                          for file in self.error_markers: 
                            total_error_number = 0
                            error_string = ''
                            for day in range(7):
                              total_error_number += len(self.error_markers[file][day])
                              error_string += f'{len(self.error_markers[file][day])}:'

                            if total_error_number == 0:
                              continue

                            button = {'type': 'button', 
                                      'dimensions': (700, 40), 
                                      'position': (5, yposition), 
                                      'bg_color': {'idle': 'black',
                                            'hover': (30, 30, 30),
                                            'clicked': 'white'}, 
                                      'bd_color': {'idle': 'white',
                                            'hover': 'black',
                                            'clicked': 'black'}, 
                                      'txt_color': {'idle': 'white',
                                            'hover': 'white',
                                            'clicked': 'black'},  
                                      'txt_font': 'button_default', 
                                      'txt_align': 'left', 
                                      'txt_content': f'{error_string} error/s | {file}', 
                                      'elements': [], 
                                      'buttonID': file}
                            yposition += 45
                            
                            window_buttons.append(button)
                          
                          self.ui_elements['windows'].append(
                            Window(self, self.display, 
                                  windowID= 'SheetListWindow',
                                  window_type= 'scrollable',
                                  dimensions= (720, 500),
                                  position= (300, 50),
                                  elements= window_buttons)
                          )

                      else:
                        if check_window(self, 'SheetListWindow'):
                          close_window(self, 'SheetListWindow')
                        else:
                          self.ui_elements['windows'].append(
                              Window(self, self.display, 
                                    windowID= 'SheetListWindow',
                                    window_type= 'scrollable',
                                    dimensions= (720, 500),
                                    position= (300, 50),
                                    elements= [{'type': 'label',
                                      'position': (10, 10),
                                      'txt_font': 'label',
                                      'txt_content': 'SCAN SHEETS FIRST',
                                      'txt_color': 'white'}])
                          )

                    case 'HowToUseButton':
                      return ['open_instructions']

                    case 'ExitButton':
                      pygame.quit()
                      sys.exit()
                    case _:
                      print('ERROR: INVALID BUTTON ID')
                      assert False
                  break			

            else:
              for window in reversed(self.ui_elements['windows'].copy()):

                for element in reversed(window.elements.copy()):
                  if element['type'] == 'button':
                    if element['content'].get_rect(offset=window.position).collidepoint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)) and element['content'].state == 'clicked':
                      match element['content'].buttonID:
                        case _:
                          if element['content'].buttonID.find('.xlsx') != -1 or element['content'].buttonID.find('.xlsm') != -1:
                            print(f"opening file: {element['content'].buttonID}")
                            return ['run_open_sheet', element['content'].buttonID, scanned_week_folder, scanned_days]
                          else:
                            print('ERROR: INVALID BUTTON ID')
                            assert False
                      break

          if event.button == 2:
            self.player_inputs['mouse'][2] = False
          if event.button == 3:
            self.player_inputs['mouse'][1] = False
    
        if event.type == pygame.QUIT:
          pygame.quit()
          sys.exit()

      self.screen.blit(self.display, (0, 0))

      pygame.display.update()
      self.clock.tick(60)

  def run_open_sheet(self, file, week_folder, scanned_days):
    def grey_out_sheet():
      for window in self.ui_elements['windows']:
        if window.windowID == 'WorksheetDisplayWindow':
          for element in window.elements:
            if element['type'] == 'input_textbox':
              print('asdasdasdasdasd')
              element['content'].bd_color = (100, 100, 100)
              print(element['content'].bd_color)
              element['content'].txt_color = (100, 100, 100)
              print(element['content'].txt_color)

    def change_sheet(sheet_number):
      nonlocal active_sheet
      print(active_sheet)
      if active_sheet != sheet_number:
        active_sheet = sheet_number

        open_sheet(sheet_number)
      
    def task_format(task_name: str):
      if task_name == None:
        return ''
      
      ## if the task name was a number, it will be considered and integer, and we would need to change its type to a string
      task_name = str(task_name)
      return task_name.upper().replace(' ', '')

    def save_sheet(sheet_number):
      def is_changed(cell):
        if  write_sheet[cell].value == None:
          return '' != extract_data(self, f'{cell}InputBox')
        return write_sheet[cell].value != extract_data(self, f'{cell}InputBox')

      write_workbook = openpyxl.load_workbook(filename=f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}/{file}', read_only= False, data_only=True)
      write_workbook.active = sheet_number
      write_sheet = write_workbook.active

      for i in range(ALLOCATION_TABLE_DIMENSIONS[0]):
        task_cell = f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}'
        
        match check_error(task_cell, sheet_number):
          case 'BgenNumberNotAdded':
            if is_changed(task_cell):  
              match self.pop_up_window(Window(self,self.display, 
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{task_cell}] Original Error: {check_error(task_cell, sheet_number)} \n
Do you want to replace the original task name: {task_format(write_sheet[task_cell].value)} \n
With: {extract_data(self, f'{task_cell}InputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'PROCEED', 
                                          'elements': [], 
                                          'buttonID': f'ProceedButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                case 'ProceedButton':
                  write_sheet[task_cell] = extract_data(self, f'{task_cell}InputBox')
                case 'CancelButton':
                  pass
                case _:
                  print("INVAILD POPUP WINDOW RESULT")
                  assert False
            
            else:
              match self.pop_up_window(Window(self,self.display, 
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{task_cell}] Original Error: {check_error(task_cell, sheet_number)} \n
Do you want to add the Bgen Number: {task_format(write_sheet[task_cell].value)} \n
to the BGEN NUMBER list?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'YES', 
                                          'elements': [], 
                                          'buttonID': f'YesButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'NO', 
                                          'elements': [], 
                                          'buttonID': f'NoButton'}
                                         ]
                                        )):
                case 'YesButton':
                  self.bgen_numbers.add(int(write_sheet[task_cell].value.strip().replace('BGENGRAM', '')))
                case 'NoButton':
                  pass
                case _:
                  print("INVAILD POPUP WINDOW RESULT")
                  assert False

          case 'InvalidTaskName':
            if is_changed(task_cell):  
              match self.pop_up_window(Window(self,self.display, 
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{task_cell}] Original Error: {check_error(task_cell, sheet_number)} \n
Do you want to replace the original task name: {task_format(write_sheet[task_cell].value)}\n
With: {extract_data(self, f'{task_cell}InputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'PROCEED', 
                                          'elements': [], 
                                          'buttonID': f'ProceedButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                
                case 'ProceedButton': ## unused feature for adding typo fixes to the auto correct
                  ## before_task_name = task_format(write_sheet[task_cell].value)
                  write_sheet[task_cell] = extract_data(self, f'{task_cell}InputBox')
                  """
                  match self.pop_up_window(Window(self,self.display, 
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{task_cell}] Original Error: {check_error(task_cell, sheet_number)} \n
Do you want to add the auto correct:  \n
{before_task_name} ---> {extract_data(self, f'{task_cell}InputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'GENERAL', 
                                          'elements': [], 
                                          'buttonID': f'GeneralButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (300, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'SPECIFIC', 
                                          'elements': [], 
                                          'buttonID': f'SpecificButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'Skip', 
                                          'elements': [], 
                                          'buttonID': f'SkipButton'}
                                         ]
                                        )):
                    case 'GeneralButton':
                      self.general_task_names_exceptions[before_task_name] = extract_data(self, f'{task_cell}InputBox')

                    case 'SpecificButton':
                      week_specific_exceptions_file = open(f'{WEEKLY_EXCEPTIONS_FILE_PATH}{week_folder}.json', 'r')
                      week_specific_task_name_exceptions = json.load(week_specific_exceptions_file)
                      week_specific_exceptions_file.close()
                    
                      week_specific_task_name_exceptions[sheet_number][before_task_name] = extract_data(self, f'{task_cell}InputBox')
                      week_specific_exceptions_file = open(f'{WEEKLY_EXCEPTIONS_FILE_PATH}{week_folder}.json', 'w')
                      json.dump(week_specific_task_name_exceptions, week_specific_exceptions_file)
                      week_specific_exceptions_file.close()

                    case 'SkipButton':
                      pass"""

                case 'CancelButton':
                  pass

                case _:
                  print("INVAILD POPUP WINDOW RESULT")
                  assert False
            
            else:
              match self.pop_up_window(Window(self,self.display, 
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{task_cell}] Original Error: {check_error(task_cell, sheet_number)}\n
Do you want to add the Task Name: {task_format(write_sheet[task_cell].value)}\n
to the Iso Numbers/Task Codes list?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'ISO NUMBERS', 
                                          'elements': [], 
                                          'buttonID': f'IsoNumbersButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (300, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'WORK CODES', 
                                          'elements': [], 
                                          'buttonID': f'WorkCodesButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                case _:
                  pass
                  # print("INVAILD POPUP WINDOW RESULT")
                  # assert False

          case 'NoError':
            if is_changed(task_cell):  
              match self.pop_up_window(Window(self,self.display, 
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{task_cell}] THERE IS NO ERROR HERE \n
Do you want to replace the original task name: {task_format(write_sheet[task_cell].value)} \n
With: {extract_data(self, f'{task_cell}InputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'PROCEED', 
                                          'elements': [], 
                                          'buttonID': f'ProceedButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                case 'ProceedButton':
                  write_sheet[task_cell] = extract_data(self, f'{task_cell}InputBox')
                case 'CancelButton':
                  pass
                case _:
                  print("INVAILD POPUP WINDOW RESULT")
                  assert False
            
            else:
              pass
          
          case _:
            print(f'INVALID ERROR IN {task_cell}')
      else: 
          pass

      ## we then handle the employee names
      ## for now, I just gotta pray to god that there will never be an occasion where 2 different names get a typo that makes them the same, 
      ## basically any 'misspelt' name is treated as another valid name and the misspelt name gets saved into the names list
      for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
        employee_name_cell = f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'
        
        match check_error(employee_name_cell, sheet_number):
          case 'InvalidEmployeeName':
            if is_changed(employee_name_cell):
              match self.pop_up_window(Window(self,self.display,
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{employee_name_cell}] Original Error: {check_error(employee_name_cell, sheet_number)}\n
Do you want to replace the original employee name: {write_sheet[employee_name_cell].value}\n
With: {extract_data(self, f'{employee_name_cell}InputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'PROCEED', 
                                          'elements': [], 
                                          'buttonID': f'ProceedButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                case _:
                  pass
                  # print("INVAILD POPUP WINDOW RESULT")
                  # assert False

            else:
              match self.pop_up_window(Window(self,self.display,
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{employee_name_cell}] Original Error: {check_error(employee_name_cell, sheet_number)}\n
Do you want to set employee name: {write_sheet[employee_name_cell].value}\n
To the code: {extract_data(self, f'{employee_name_cell}CodeInputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'PROCEED', 
                                          'elements': [], 
                                          'buttonID': f'ProceedButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                case _:
                  pass
                  # print("INVAILD POPUP WINDOW RESULT")
                  # assert False
                                      
          case 'NoError':
            if is_changed(employee_name_cell):
              match self.pop_up_window(Window(self,self.display,
                                         windowID= 'SaveSheetPopUpWindow',
                                         window_type= 'focused',
                                         dimensions= (800, 600),
                                         position= (560, 240),
                                         elements= [
                                          {'type': 'textbox_big', 
                                          'text': f'''[{employee_name_cell}] THERE IS NO ERROR HERE\n
Do you want to replace the original employee name: {write_sheet[employee_name_cell].value}\n
With: {extract_data(self, f'{employee_name_cell}InputBox')}   ?''',
                                          'color': 'white', 
                                          'dimensions': (700, 400), 
                                          'position': (50, 50)},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (50, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'PROCEED', 
                                          'elements': [], 
                                          'buttonID': f'ProceedButton'},
                                          {'type': 'button', 
                                          'dimensions': (200, 40), 
                                          'position': (550, 530), 
                                          'bg_color': {'idle': 'black',
                                                      'hover': (30, 30, 30),
                                                      'clicked': 'white'}, 
                                          'bd_color': {'idle': 'white',
                                                      'hover': 'black',
                                                      'clicked': 'black'}, 
                                          'txt_color': {'idle': 'white',
                                                      'hover': 'white',
                                                      'clicked': 'black'},  
                                          'txt_font': 'button_default', 
                                          'txt_align': 'center', 
                                          'txt_content': f'CANCEL', 
                                          'elements': [], 
                                          'buttonID': f'CancelButton'}
                                         ]
                                        )):
                case _:
                  pass
                  # print("INVAILD POPUP WINDOW RESULT")
                  # assert False
            
            else:
              pass

      for i in range(ALLOCATION_TABLE_DIMENSIONS[0]):
        for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
          hour_cell = f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{ALLOCATION_TABLE_START_CORNER[1] + j}'
          write_sheet[hour_cell] = extract_data(self, f'{hour_cell}InputBox')
      
      write_workbook.save(filename=f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}/{file}')
      saved_days[sheet_number] = 1
      self.scan_needed = True
      print('asd')
      grey_out_sheet()


    def open_sheet(sheet_number):
      top_bar_buttons = []

      xposition = 5
      for i in range(7):
        if scanned_days[i] == 1:
          if i == sheet_number:
            if len(self.error_markers[file][i]) == 0:
              button_bg_color = active_sheet_button_colors[0]['bg_color']
              button_bd_color = active_sheet_button_colors[0]['bd_color']
              button_txt_color = active_sheet_button_colors[0]['txt_color']
            else: 
              button_bg_color = active_sheet_button_colors[1]['bg_color']
              button_bd_color = active_sheet_button_colors[1]['bd_color']
              button_txt_color = active_sheet_button_colors[1]['txt_color']

          else:
            if len(self.error_markers[file][i]) == 0:
              button_bg_color = inactive_sheet_button_colors[0]['bg_color']
              button_bd_color = inactive_sheet_button_colors[0]['bd_color']
              button_txt_color = inactive_sheet_button_colors[0]['txt_color']
            else: 
              button_bg_color = inactive_sheet_button_colors[1]['bg_color']
              button_bd_color = inactive_sheet_button_colors[1]['bd_color']
              button_txt_color = inactive_sheet_button_colors[1]['txt_color']

          button = {'type': 'button', 
                    'dimensions': (100, 40), 
                    'position': (xposition, 5), 
                    'bg_color': button_bg_color, 
                    'bd_color': button_bd_color, 
                    'txt_color': button_txt_color,  
                    'txt_font': 'button_default', 
                    'txt_align': 'left', 
                    'txt_content': f'{weekdays[i]}  [{len(self.error_markers[file][i])}]', 
                    'elements': [], 
                    'buttonID': f'{weekdays[i]}SheetButton'}

          top_bar_buttons.append(button)
        
        xposition += 105

      read_workbook = openpyxl.load_workbook(filename=f'{ALLOCATION_TABLE_FILE_PATH}{week_folder}/{file}', data_only=True)

      screen_input_boxes = load_sheet(read_workbook, sheet_number)

      read_workbook.close()

      self.ui_elements = {
        'windows': [Window(self, self.display, 
                          windowID= "DaySelectionWindow", 
                          dimensions= (800, 50), position= (10, 10), 
                          elements= top_bar_buttons),
                    Window(self, self.display,
                          windowID="WorksheetDisplayWindow",
                          dimensions= (1900, 1000), position= (10, 70),
                          elements= screen_input_boxes),
                    Window(self, self.display,
                          windowID="WorkBookFileNameWindow",
                          dimensions= (600, 50), position= (820, 10),
                          elements= [{'type': 'label',
                                    'position': (10, 10),
                                    'txt_font': 'label',
                                    'txt_content': file,
                                    'txt_color': 'white'}])],

         

        'buttons': [Button(self, self.display,
                            dimensions=(100, 40),
                            position= (1810, 15),
                            txt_content= 'BACK',
                            buttonID= 'BackButton'),
                    Button(self, self.display,
                            dimensions=(100, 40),
                            position= (1700, 15),
                            txt_content= 'SAVE',
                            buttonID= 'SaveButton')]
      }

      if saved_days[sheet_number] == 1:
        grey_out_sheet()

    def check_error(cell, sheet_number):
      for error in self.error_markers[file][sheet_number]:
        if error[0] == cell:
          return error[1]
        
      return 'NoError'

    def load_sheet(read_workbook, sheet_number):
      read_workbook.active = sheet_number
      read_sheet = read_workbook.active
      
      worksheet_input_boxes = []
      screen_pos = [0, 0]

      ## we first fill in the task names
      screen_pos = [465, 5]

      for i in range(ALLOCATION_TABLE_DIMENSIONS[0]):
        task_cell = f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{TASK_NAME_ROW}'
        print(task_cell)
        if read_sheet[task_cell].value == None:  
          cell_content = ''
        else:  
          cell_content = str(read_sheet[task_cell].value)
        
        match check_error(task_cell, sheet_number):
          case 'BgenNumberNotAdded':
            worksheet_input_boxes.append({'type': 'input_textbox/vert_colored',
                                     'dimensions': (50, 300),
                                     'position': tuple(screen_pos),
                                     'bd_color': 'orange',
                                     'txt_color': 'orange',
                                     'txt_content': cell_content,
                                     'input_boxID': f'{task_cell}InputBox',
                                     'char_limit': None})

          case 'InvalidTaskName':
            worksheet_input_boxes.append({'type': 'input_textbox/vert_colored',
                                     'dimensions': (50, 300),
                                     'position': tuple(screen_pos),
                                     'bd_color': 'red',
                                     'txt_color': 'red',
                                     'txt_content': cell_content,
                                     'input_boxID': f'{task_cell}InputBox',
                                     'char_limit': None})

          case 'NoError':
            worksheet_input_boxes.append({'type': 'input_textbox/vert',
                                      'dimensions': (50, 300),
                                      'position': tuple(screen_pos),
                                      'txt_content': cell_content,
                                      'input_boxID': f'{task_cell}InputBox',
                                      'char_limit': None})
            
          case _:
            print(f'INVALID ERROR IN {task_cell}')

        screen_pos[0] += 55

      ## we then fill in the employee names
      screen_pos = [5, 310]
      
      for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
        employee_name_cell = f'{EMPLOYEE_NAME_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'
        if read_sheet[employee_name_cell].value == None:  
          cell_content = ''
        else:  
          cell_content = str(read_sheet[employee_name_cell].value)
        
        match check_error(employee_name_cell, sheet_number):
          case 'InvalidEmployeeName':
            worksheet_input_boxes.append({'type': 'input_textbox/hori_colored',
                                     'dimensions': (310, 40),
                                     'position': tuple(screen_pos),
                                     'bd_color': 'red',
                                     'txt_color': 'red',
                                     'txt_content': cell_content,
                                     'input_boxID': f'{employee_name_cell}InputBox',
                                     'char_limit': None})
            screen_pos[0] += 315

            worksheet_input_boxes.append({'type': 'input_textbox/hori_colored',
                                     'dimensions': (140, 40),
                                     'position': tuple(screen_pos),
                                     'bd_color': 'red',
                                     'txt_color': 'red',
                                     'txt_content': '',
                                     'input_boxID': f'{employee_name_cell}CodeInputBox',
                                     'char_limit': None})
            screen_pos[0] = 5

          case 'NoError':
            worksheet_input_boxes.append({'type': 'input_textbox/default',
                                     'dimensions': (310, 40),
                                     'position': tuple(screen_pos),
                                     'txt_content': cell_content,
                                     'input_boxID': f'{employee_name_cell}InputBox',
                                     'char_limit': None})
            screen_pos[0] += 315

            worksheet_input_boxes.append({'type': 'input_textbox/hori_colored',
                                     'dimensions': (140, 40),
                                     'position': tuple(screen_pos),
                                     'bd_color': 'white',
                                     'txt_color': 'white',
                                     'txt_content': self.employee_names[cell_content.strip()] if cell_content != None and cell_content.strip() != '' else '',
                                     'input_boxID': f'{employee_name_cell}CodeInputBox',
                                     'char_limit': None})
            screen_pos[0] = 5
        
        screen_pos[1] += 45

      ## we then fill in the employee roles
      screen_pos = [260, 310]

      """
      for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
        employee_role_cell = f'{EMPLOYEE_ROLE_COLUMN}{ALLOCATION_TABLE_START_CORNER[1] + j}'
        if read_sheet[employee_role_cell].value == None:  
          cell_content = ''
        else:  
          cell_content = str(read_sheet[employee_role_cell].value)

        worksheet_input_boxes.append({'type': 'input_textbox/default',
                                     'dimensions': (200, 40),
                                     'position': tuple(screen_pos),
                                     'txt_content': cell_content,
                                     'input_boxID': f'{employee_role_cell}InputBox',
                                     'char_limit': None})
          
        screen_pos[1] += 45
      """

      ## finally we fill in the hours
      screen_pos = [465, 310]

      for i in range(ALLOCATION_TABLE_DIMENSIONS[0]):
        for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
          hour_cell = f'{alphabet_converter(ALLOCATION_TABLE_START_CORNER[0] + i)}{ALLOCATION_TABLE_START_CORNER[1] + j}'
          if read_sheet[hour_cell].value == None:  
            cell_content = ''
          else:  
            cell_content = str(read_sheet[hour_cell].value)

          print(cell_content)

          match check_error(hour_cell, sheet_number):
            case 'EmployeeNameMissing':
              worksheet_input_boxes.append({'type': 'input_textbox/small',
                                      'dimensions': (50, 40),
                                      'position': tuple(screen_pos),
                                      'bd_color': 'orange', 
                                      'txt_color': 'orange',
                                      'txt_content': cell_content,
                                      'input_boxID': f'{hour_cell}InputBox',
                                      'char_limit': None})
              
            case 'TaskNameMissing':
              worksheet_input_boxes.append({'type': 'input_textbox/small',
                                      'dimensions': (50, 40),
                                      'position': tuple(screen_pos),
                                      'bd_color': (191, 71, 255), 
                                      'txt_color': (191, 71, 255),
                                      'txt_content': cell_content,
                                      'input_boxID': f'{hour_cell}InputBox',
                                      'char_limit': None})

            case 'WhitespaceInCell':
              worksheet_input_boxes.append({'type': 'input_textbox/small',
                                      'dimensions': (50, 40),
                                      'position': tuple(screen_pos),
                                      'bd_color': 'red', 
                                      'txt_color': 'red',
                                      'txt_content': cell_content,
                                      'input_boxID': f'{hour_cell}InputBox',
                                      'char_limit': None})

            case _:
              worksheet_input_boxes.append({'type': 'input_textbox/small',
                                      'dimensions': (50, 40),
                                      'position': tuple(screen_pos),
                                      'bd_color': 'white', 
                                      'txt_color': 'white',
                                      'txt_content': cell_content,
                                      'input_boxID': f'{hour_cell}InputBox',
                                      'char_limit': None})
          
          screen_pos[1] += 45
        screen_pos[0] += 55
        screen_pos[1] = 310
      return worksheet_input_boxes

    self.display = pygame.Surface((1920, 1080))

    saved_days = [0, 0, 0, 0, 0, 0, 0]

    inactive_sheet_button_colors = [{'bg_color': {'idle': 'black',
                                          'hover': (30, 30, 30),
                                          'clicked': 'white'}, 
                                    'bd_color': {'idle': 'white',
                                          'hover': 'white',
                                          'clicked': 'black'}, 
                                    'txt_color': {'idle': 'white',
                                          'hover': 'white',
                                          'clicked': 'black'}},
                                    {'bg_color': {'idle': 'black',
                                          'hover': (30, 30, 30),
                                          'clicked': 'white'}, 
                                    'bd_color': {'idle': 'red',
                                          'hover': 'red',
                                          'clicked': 'black'}, 
                                    'txt_color': {'idle': 'red',
                                          'hover': 'red',
                                          'clicked': 'red'}}]
    active_sheet_button_colors = [{'bg_color': {'idle': 'blue',
                                          'hover': 'cyan',
                                          'clicked': 'white'}, 
                                    'bd_color': {'idle': 'white',
                                          'hover': 'white',
                                          'clicked': 'black'}, 
                                    'txt_color': {'idle': 'white',
                                          'hover': 'white',
                                          'clicked': 'black'}},
                                    {'bg_color': {'idle': 'blue',
                                          'hover': 'cyan',
                                          'clicked': 'white'}, 
                                    'bd_color': {'idle': 'red',
                                          'hover': 'red',
                                          'clicked': 'black'}, 
                                    'txt_color': {'idle': 'red',
                                          'hover': 'red',
                                          'clicked': 'red'}}]
    
    weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    
    active_sheet = None

    initial_sheet_number = 0
    for day in scanned_days:
      if day == 0:
        initial_sheet_number += 1
      else:
        break
    
    print(initial_sheet_number)
    active_sheet = initial_sheet_number

    open_sheet(initial_sheet_number)

    mouse_position = (0, 0)

    while True:
      self.display.fill((0, 0, 0))

      ## sprint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock))
      mouse_position = scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)

      mouse_on_window = update_ui_elements(self)

      			## RENDERING SECTION ##

      render_ui(self)

      pygame.draw.circle(self.display, 'red', mouse_position, 5, 1)

      ##### EVENT HANDLING SECTION #####	

      for event in pygame.event.get():

        if event.type == pygame.KEYDOWN:
          ui_process_keyboard_button_down(self, event)

        if event.type == pygame.MOUSEBUTTONDOWN:
          
          if event.button == 1:
            self.player_inputs['mouse'][0] = True
            ui_event = ui_process_mouse_button_down(self, 1, mouse_on_window)
            if ui_event != None:
              match ui_event['type']:

                case 'textbox_conclusion':

                  for action in ui_event['actions']:	

                    match action[0]:

                      case 'remove_window':
                        
                        self.ui_elements['windows'].remove(action[1])
          if event.button == 2:
            self.player_inputs['mouse'][2] = True
          if event.button == 3:
            self.player_inputs['mouse'][1] = True

        if event.type == pygame.MOUSEWHEEL:
          print(event.y)
          if event.y > 0:
            ui_process_mouse_button_down(self, 4, mouse_on_window)  
          elif event.y < 0:
            ui_process_mouse_button_down(self, 5, mouse_on_window)  

        if event.type == pygame.MOUSEBUTTONUP:
          if event.button == 1:
            self.player_inputs['mouse'][0] = False

            ### I think I'm gonna not do a function for this, it just feels too specialised

            if not mouse_on_window:
              for button in reversed(self.ui_elements['buttons'].copy()):
                if button.get_rect().collidepoint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)) and button.state == 'clicked':
                  match button.buttonID:
                    case 'BackButton':
                      return ['start_screen', [week_folder[3:5], week_folder[6:8], scanned_days, self.error_markers.copy()]]
                    case 'SaveButton':
                      if saved_days[active_sheet] == 0:
                        save_sheet(active_sheet)
                      
                    case _:
                      print('ERROR: INVALID BUTTON ID')
                      assert False
                  break			

            else:
              for window in reversed(self.ui_elements['windows'].copy()):

                for element in reversed(window.elements.copy()):
                  if element['type'] == 'button':
                    if element['content'].get_rect(offset=window.position).collidepoint(scale_pos(self, pygame.mouse.get_pos(), self.scale_lock)) and element['content'].state == 'clicked':
                      match element['content'].buttonID:
                        case 'MonSheetButton':  
                          change_sheet(0)
                        
                        case 'TueSheetButton':  
                          change_sheet(1)

                        case 'WedSheetButton':  
                          change_sheet(2)

                        case 'ThuSheetButton':  
                          change_sheet(3)

                        case 'FriSheetButton':  
                          change_sheet(4)

                        case 'SatSheetButton':  
                          change_sheet(5)

                        case 'SunSheetButton':  
                          change_sheet(6)

                        case _:
                        
                          print('ERROR: INVALID BUTTON ID')
                          assert False
                      break

          if event.button == 2:
            self.player_inputs['mouse'][2] = False
          if event.button == 3:
            self.player_inputs['mouse'][1] = False
    
        if event.type == pygame.QUIT:
          pygame.quit()
          sys.exit()

      self.screen.blit(pygame.transform.scale(self.display, (self.screen.get_width(), self.display.get_height()*self.screen.get_width() / self.display.get_width())), (0, 0))
      ## print(random.random())
      pygame.display.update()
      self.clock.tick(20)

Application().run_engine()