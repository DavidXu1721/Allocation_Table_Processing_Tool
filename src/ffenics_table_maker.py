import openpyxl
import os
import time
import json
from textwrap import dedent, shorten, wrap

CURRENTYEAR = 25 # Make sure to change this every year
SCREENWIDTH = 80
ALLOCATION_TABLE_DIMENSIONS = (26, 12)

NAME_FILE_PATH = "config/employee_names.json"
CODE_FILE_PATH = "config/code_to_name.json"
GENERAL_EXCEPTIONS_FILE_PATH = "config/general_exceptions.json"
WEEKLY_EXCEPTIONS_FILE_PATH = 'config/weekly_exceptions/'
VARIATION_NUMBERS_FILE_PATH = "config/variation_numbers.json"
ISO_NUMBERS_FILE_PATH = "config/iso_numbers.json"
ALLOCATION_TABLE_FILE_PATH = "Allocation_Tables/"

name_file = open(NAME_FILE_PATH, 'r')
NAMES = json.load(name_file)
name_file.close()

code_file = open(CODE_FILE_PATH, 'r')
CODES = json.load(code_file)
code_file.close()

bgen_numbers_file = open(VARIATION_NUMBERS_FILE_PATH, 'r')
BGEN_NUMBERS = set(json.load(bgen_numbers_file))
bgen_numbers_file.close()
## print(self.bgen_numbers)


iso_numbers_file = open(ISO_NUMBERS_FILE_PATH, 'r')
ISOS = set(json.load(iso_numbers_file))
iso_numbers_file.close()

TASK_NAME_EXCEPTIONS ={
	
}

employees_not_present = []
total_employee_names_used = []

employee_absences_list = []

def textbox(text):
	print(dedent(f"""\
		{'#' * SCREENWIDTH}"""))
	for line in wrap(text, SCREENWIDTH - 4):
		print(dedent(f"""\
		# {line}{' ' * (SCREENWIDTH - 4 - len(line))} #"""))
	print(dedent(f"""\
		{'#' * SCREENWIDTH}"""))
	
def alphabet_converter(n):
	"""Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
	name = ''
	while n > 0:
		n, r = divmod (n - 1, 26)
		name = chr(r + ord('A')) + name
	return name

def log_employee(employee_name):
	if employee_name in total_employee_names_used:
		print(f"ERROR: Same name {employee_name} appears in two sheets!")
		exit(1)
	else:
		if employee_name not in employee_names_used:
			employee_names_used.append(employee_name)

def read_cells(row, start_column, length, read_sheet):
	result = ''
	i = 0
	while i < length:
		if read_sheet[f'{alphabet_converter(4 + start_column + i)}{9 + row}'].value == None:
			char = ''
		else:
			char = str(read_sheet[f'{alphabet_converter(4 + start_column + i)}{9 + row}'].value)
		result = result + char
		i += 1
	return result

def add_absence(namecode, day_num, activity): ## day_num is 0-6, not 1-7
	entry = [activity, None, namecode, None]

	days = [None for i in range(7)]
	days[day_num] = 0

	entry.extend(days)

	employee_absences_list.append(entry)

	

textbox(f"""Year: 20{CURRENTYEAR}
		Welcome to this ffenics helper, by David Xu
		Make sure that all the relevant workbooks are of the file type .xlsx
		IMPORTANT: make sure that all allocation sheets are closed!!!
		PS: Make sure that the year is right .""")

month = input("INPUT MONTH Folder: ")
week_end = input("INPUT WEEK END date: ")
days_specified = input("INPUT DAY (leave blank to get all days): ")
skip_to_table = input("Skip to table? (Leave blank to not skip): ")
show_absences = input("Show absences? (Leave blank to disable): ")


if days_specified == '':
	sheet_range = range(7)
else:
	sheet_range = range(int(days_specified))

if skip_to_table == '':
	write_workbook = openpyxl.Workbook()

	for file in sorted(os.listdir(f"Allocation_Tables/WE.{week_end}.{month}.{CURRENTYEAR}")):
		# The program first works through each individual file
		read_workbook = openpyxl.load_workbook(filename=f'Allocation_Tables/WE.{week_end}.{month}.{CURRENTYEAR}/{file}', read_only=False, data_only=True)
		
		employee_names_used = []
		for sheet_number in sheet_range:
			# And then through each individual sheet
			read_workbook.active = sheet_number
			print("\n" + file + str(sheet_number) + "~###################################################################################################################################################~~")
			read_sheet = read_workbook.active

			if len(read_workbook.sheetnames) != 9: ## tbh, I don't know why I made this run in every sheet
				print(read_workbook.sheetnames)
				print("Error, Extra Sheet")
				exit(1)            

			for i in range(ALLOCATION_TABLE_DIMENSIONS[0]):
				for j in range(ALLOCATION_TABLE_DIMENSIONS[1]):
					current_cell = read_sheet[f'{alphabet_converter(4 + i)}{9 + j}'].value
					print(f'{alphabet_converter(4 + i)}{9 + j}', end=' ')
					### print(f'{alphabet_converter(4 + i)}{9 + j} --- {current_cell}')

					# this checks if the current cell is a number and that the corresponding employee is not a supervisor.
					if current_cell != None and str(current_cell).replace('.', '').isnumeric() and not '(' in NAMES[read_sheet[f"B{9 + j}"].value.strip()]:	
						employee_name = read_sheet[f"B{9 + j}"].value
						log_employee(NAMES[employee_name.strip()])
						task: str = read_sheet[f"{alphabet_converter(4 + i)}3"].value
						print(task, end= " ")
						print(f'{employee_name} did {task.upper().replace(" ", "")} for {current_cell} hours')
						
						if task.upper().replace(' ', '') in write_workbook.sheetnames:
							write_workbook.active = write_workbook[task.upper().replace(' ', '')]
							write_sheet = write_workbook.active

						elif task != None:

							if task.upper().replace(' ', '') in TASK_NAME_EXCEPTIONS.keys():
								task_exception = TASK_NAME_EXCEPTIONS[task.upper().replace(' ', '')]
								print(f'{task} replaced with {task_exception}')
								task = task_exception
								read_sheet[f"{alphabet_converter(4 + i)}3"] = task.upper().replace(' ', '')

							elif not (''.join([i for i in task.upper().replace(' ', '') if not i.isdigit()]) in ['ZZ-', 'SI-', 'BGENGRAM']) and not (task.upper().replace(' ', '') in ISOS):
								task_exception = input(f"Error: {task.upper().replace(' ', '')}")
								if task_exception != '':
									print(f"{task.upper().replace(' ', '')} replaced with {task_exception}")
									task = task_exception
									read_sheet[f"{alphabet_converter(4 + i)}3"] = task.upper().replace(' ', '')

							## This code writes into the summary sheet
							
							if task.upper().replace(' ', '') in write_workbook.sheetnames:
								write_workbook.active = write_workbook[task.upper().replace(' ', '')]
								write_sheet = write_workbook.active

							else:
								write_workbook.create_sheet(task.upper().replace(' ', ''))
								### print(write_workbook.sheetnames)
								write_workbook.active = write_workbook[task.upper().replace(' ', '')]
								write_sheet = write_workbook.active
								if 'BGENGRAM' in task.upper().replace(' ', ''):
									if not int(task.upper().replace(' ', '').replace('BGENGRAM', '')) in BGEN_NUMBERS:
										input(f"{task.upper().replace(' ', '')} is not added yet.")
									write_sheet['A1'] = task.upper().replace(' ', '').replace('BGENGRAM', '')
								else:
									write_sheet['A1'] = task.upper().replace(' ', '')
								write_sheet['B1'] = 'Mon'
								write_sheet['C1'] = 'Tues'
								write_sheet['D1'] = 'Wed'
								write_sheet['E1'] = 'Thur'
								write_sheet['F1'] = 'Fri'
								write_sheet['G1'] = 'Sat'
								write_sheet['H1'] = 'Sun'
						else:
							error_input = ''
							while error_input != '1':
								error_input = input('ERROR! Task cell is blank! Enter 1 to continue:')
								
						row_index = 2
						while write_sheet[f'A{row_index}'].value != NAMES[employee_name.strip()] and write_sheet[f'A{row_index}'].value != None:
							row_index += 1
						write_sheet[f'A{row_index}'] = NAMES[employee_name.strip()]
						if write_sheet[f'{alphabet_converter(sheet_number + 2)}{row_index}'].value == None:
							write_sheet[f'{alphabet_converter(sheet_number + 2)}{row_index}'] = current_cell
						else:
							input("ERROR: there's an hour value already here!")

						
					elif str(read_sheet[f"{alphabet_converter(4 + i)}3"].value).replace(' ', '') == 'ZZ-007' and read_sheet[f"B{9 + j}"].value != None and not '(' in NAMES[read_sheet[f"B{9 + j}"].value.strip()]:
						employee_name = read_sheet[f"B{9 + j}"].value
						log_employee(NAMES[employee_name.strip()])
						row_index = 2
						### print(f'Adding {employee_name} for ZZ-007')
						
						if 'ZZ-007' in write_workbook.sheetnames:
							write_workbook.active = write_workbook['ZZ-007']
							write_sheet = write_workbook.active
						else:
							write_workbook.create_sheet('ZZ-007')
							write_workbook.active = write_workbook['ZZ-007']
							write_sheet = write_workbook.active
							write_sheet['A1'] = 'ZZ-007'
							write_sheet['B1'] = 'Mon'
							write_sheet['C1'] = 'Tues'
							write_sheet['D1'] = 'Wed'
							write_sheet['E1'] = 'Thur'
							write_sheet['F1'] = 'Fri'
							write_sheet['G1'] = 'Sat'
							write_sheet['H1'] = 'Sun'
							
						while write_sheet[f'A{row_index}'].value != NAMES[employee_name.strip()] and write_sheet[f'A{row_index}'].value != None:
							row_index += 1
						write_sheet[f'A{row_index}'] = NAMES[employee_name.strip()]
						if write_sheet[f'{alphabet_converter(sheet_number + 2)}{row_index}'].value == None:
							write_sheet[f'{alphabet_converter(sheet_number + 2)}{row_index}'] = current_cell
						else:
							input("ERROR: there's an hour value already here!")

					elif current_cell != None and not '(' in NAMES[read_sheet[f"B{9 + j}"].value.strip()] and not read_sheet[f"B{9 + j}"].value in employees_not_present:
						employee_name = read_sheet[f"B{9 + j}"].value
						log_employee(NAMES[employee_name.strip()])
						employees_not_present.append(employee_name)
						if show_absences != '':
							apply_to_ffenics = input(f'~~~~~~~Adding {CODES[NAMES[employee_name.strip()]]} as {read_cells(j, i, ALLOCATION_TABLE_DIMENSIONS[0] - i + 1, read_sheet)}')
							if apply_to_ffenics != '':
								add_absence(NAMES[employee_name.strip()], sheet_number, apply_to_ffenics)

			employees_not_present = []

		total_employee_names_used = total_employee_names_used + employee_names_used
		employee_names_used = []
		print(total_employee_names_used)
		
		read_workbook.save(filename=f'Allocation_Tables/WE.{week_end}.{month}.{CURRENTYEAR}/{file}')
		read_workbook.close()

	write_workbook.save(filename=f'Worked_Hours_Summary/Summary_sheet_{week_end}.{month}.{CURRENTYEAR}.xlsx')
	write_workbook.close()

prompt = ''
while prompt != '1':
	prompt = input('Ready to fill in the Tables press 1 to continue:')

summary_table = openpyxl.load_workbook(filename=f'Worked_Hours_Summary/Summary_sheet_{week_end}.{month}.{CURRENTYEAR}.xlsx', read_only=True, data_only=True)
# final_table = openpyxl.load_workbook(filename=f'Final_sheet.xlsx', read_only=False, data_only=True)
final_table = openpyxl.Workbook()
final_table_write_sheet = final_table.active

final_table_row_index = 1
final_table_column_index = 1

for entry in employee_absences_list:
	for cell in entry:
		final_table_write_sheet[f'{alphabet_converter(final_table_column_index)}{final_table_row_index}']= cell
		final_table_column_index += 1

	final_table_write_sheet[f'{alphabet_converter(final_table_column_index)}{final_table_row_index}'] = CODES[entry[2]]
	
	final_table_row_index +=1
	final_table_column_index = 1


for i in range(1, len(summary_table.sheetnames)):
	summary_table.active = i
	summary_read_sheet = summary_table.active
	print(f'Current Sheet: {i}/{len(summary_table.sheetnames)}| {summary_table.active.title}')

	activity_name = summary_read_sheet["A1"].value
	summary_row_index = 2

	while summary_read_sheet[f"A{summary_row_index}"].value != None and summary_read_sheet[f"A{summary_row_index}"].value != "":
		name_code = summary_read_sheet[f"A{summary_row_index}"].value
		print(f'EmployeeNameCode: {name_code}', end= ' ')

		final_table_write_sheet[f'{alphabet_converter(final_table_column_index)}{final_table_row_index}'] = activity_name
		final_table_column_index += 2
		final_table_write_sheet[f'{alphabet_converter(final_table_column_index)}{final_table_row_index}'] = name_code
		final_table_column_index += 2
		
		for j in range(7):
			final_table_write_sheet[f'{alphabet_converter(final_table_column_index)}{final_table_row_index}'] = summary_read_sheet[f"{alphabet_converter(2 + j)}{summary_row_index}"].value

			final_table_column_index += 1

		final_table_write_sheet[f'{alphabet_converter(final_table_column_index)}{final_table_row_index}'] = CODES[name_code]

		final_table_row_index +=1
		final_table_column_index = 1

		summary_row_index += 1
		print(' ')
	
	summary_row_index = 1

final_table.save(filename=f'Final_sheet.xlsx')

final_table.close()
summary_table.close()



